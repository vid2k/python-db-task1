# using pandas and openpyxl for converting workbook sheets to different csv files
import pandas as pd
from openpyxl import load_workbook
import os
from pathlib import Path

# loading each sheet of workbook into different dataframe with dynamic name
workbook_dir_name = "workbooks"
path_to_workbook_dir = Path(f"./{workbook_dir_name}")
# grabbing names of all .xlsx files inside workbook directory
file_list = [str(f) for f in path_to_workbook_dir.glob('**/*') if f.is_file()]
# print(file_list[0][len(workbook_dir_name)+1:])
# new directory made for each file in workbooks dir to save csv files
for file in file_list:
    file_name_with_ext = file[len(workbook_dir_name)+1:]
    file_name = file_name_with_ext[:-5]
    dir_name = f'{workbook_dir_name}/{file[len(workbook_dir_name)+1:-5]}'
    # print(dir_name)
    # print(file)
    # print(path_to_workbook_dir)
    # # print(f"{dir_name}.xlsx")
    # print(file_name_with_ext)
    # print(file_name)
    # print(file[len(workbook_dir_name)+1:-5])
    wb = load_workbook(f"./{dir_name}.xlsx")
    all_sheets = wb.sheetnames
    i = 0
    # directory where file is saved
    if not os.path.exists(f"csv_files/{file_name}"):
        os.mkdir(f"csv_files/{file_name}")
    for sheet in all_sheets:
        i += 1
        # dynamically create sheets with dynamic variable names
        # varname = "dataset" + str(i)
        varname = f"{file_name}_{sheet}"
        print(f"saving {sheet} sheet as '{file_name}_{sheet}'.csv")
        columns = next(wb[sheet].values)[0:]
        globals()[varname] = pd.DataFrame(wb[sheet].values, columns=columns) #df
        # now to remove first line of indexes:
        globals()[varname] = globals()[varname].iloc[1:]
    #     converting each to a different csv file since csv cannot contain multiple sheets
    #     path_to_folder = "./csv_files"
        globals()[varname].to_csv(f"csv_files/{file_name}/{varname}.csv", encoding='utf-8', index=False)

#     #     2 diff se import
#     #     mkdir using sys and new folder for each excel
# #     use creativity b4 after your first approach but use it later
#
